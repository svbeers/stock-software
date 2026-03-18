from __future__ import annotations

import csv
import io
import os
import re
import shutil
import sqlite3
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, Response, abort, flash, g, redirect, render_template, request, session, url_for
from werkzeug.middleware.proxy_fix import ProxyFix


BASE_DIR = Path(__file__).resolve().parent
DATABASE_PATH = Path(os.environ.get("STOCK_MANAGER_DB", str(BASE_DIR / "stock_manager.db")))
BACKUP_DIR = Path(os.environ.get("STOCK_MANAGER_BACKUP_DIR", str(DATABASE_PATH.parent / "backups")))
ADMIN_USERNAME = os.environ.get("STOCK_MANAGER_ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("STOCK_MANAGER_ADMIN_PASSWORD", "").strip()
SECRET_KEY = os.environ.get("STOCK_MANAGER_SECRET_KEY", "dev")

app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("STOCK_MANAGER_COOKIE_SECURE", "0") == "1"
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

DEFAULT_SETTINGS = {
    "business_name": "Werkvoorraad",
    "business_tagline": "Industriel voorraadbeheer voor elektrische werken",
    "article_number_label": "Artikelnummer",
    "barcode_label": "Barcode",
    "description_label": "Omschrijving",
    "unit_label": "Verkoopeenheid",
    "purchase_quantity_label": "Aantal per aankoop",
    "purchase_price_label": "Aankoopprijs",
    "stock_quantity_label": "Huidige voorraad",
    "profit_margin_label": "Winstmarge (%)",
    "category_label": "Categorie",
    "product_categories": "Kabel\nBuis\nAutomaat\nBevestiging",
}

LEGACY_DEFAULT_SETTINGS = {
    "business_name": "Workshop Stock",
    "business_tagline": "Industrial stock control for electrician jobs",
    "article_number_label": "Article number",
    "barcode_label": "Barcode",
    "description_label": "Description",
    "unit_label": "Unit you sell",
    "purchase_quantity_label": "Units in one purchase",
    "purchase_price_label": "Purchase price",
    "stock_quantity_label": "Current stock",
    "profit_margin_label": "Profit margin (%)",
    "category_label": "Category",
    "product_categories": "Cable\nTube\nBreaker\nFixings",
}


def is_auth_enabled() -> bool:
    return bool(ADMIN_PASSWORD)


def is_logged_in() -> bool:
    if not is_auth_enabled():
        return True
    return bool(session.get("authenticated"))


def login_required(view: Any) -> Any:
    @wraps(view)
    def wrapped_view(*args: Any, **kwargs: Any) -> Any:
        if not is_logged_in():
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)

    return wrapped_view


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        connection = sqlite3.connect(DATABASE_PATH)
        connection.row_factory = sqlite3.Row
        g.db = connection
    return g.db


@app.teardown_appcontext
def close_db(_: Any) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.before_request
def require_login() -> Response | None:
    if not is_auth_enabled():
        return None
    allowed_endpoints = {"login", "logout", "static"}
    if request.endpoint in allowed_endpoints:
        return None
    if not is_logged_in():
        return redirect(url_for("login", next=request.path))
    return None


def init_db() -> None:
    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(DATABASE_PATH)
    db.row_factory = sqlite3.Row
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            article_number TEXT NOT NULL,
            barcode TEXT NOT NULL DEFAULT '',
            image_url TEXT NOT NULL DEFAULT '',
            description TEXT NOT NULL,
            unit TEXT NOT NULL,
            purchase_quantity REAL NOT NULL DEFAULT 1,
            purchase_price REAL NOT NULL DEFAULT 0,
            stock_quantity REAL NOT NULL DEFAULT 0,
            cost REAL NOT NULL DEFAULT 0,
            profit_margin REAL NOT NULL DEFAULT 0,
            meter_tracking_enabled INTEGER NOT NULL DEFAULT 0,
            category TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            client_name TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            is_archived INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS job_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity REAL NOT NULL,
            article_number_snapshot TEXT NOT NULL,
            description_snapshot TEXT NOT NULL,
            unit_snapshot TEXT NOT NULL,
            unit_cost_snapshot REAL NOT NULL,
            profit_margin_snapshot REAL NOT NULL,
            is_invoiced INTEGER NOT NULL DEFAULT 0,
            invoice_number TEXT NOT NULL DEFAULT '',
            meter_start_snapshot REAL,
            meter_end_snapshot REAL,
            FOREIGN KEY (job_id) REFERENCES jobs (id) ON DELETE CASCADE,
            FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE RESTRICT
        );

        CREATE TABLE IF NOT EXISTS stock_purchases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            packages_received REAL NOT NULL,
            units_added REAL NOT NULL,
            package_price_snapshot REAL NOT NULL,
            unit_snapshot TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS stock_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            quantity_remaining REAL NOT NULL,
            unit_cost REAL NOT NULL,
            source_type TEXT NOT NULL DEFAULT 'manual',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS job_material_batch_allocations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_material_id INTEGER NOT NULL,
            batch_id INTEGER,
            quantity REAL NOT NULL,
            unit_cost_snapshot REAL NOT NULL,
            FOREIGN KEY (job_material_id) REFERENCES job_materials (id) ON DELETE CASCADE,
            FOREIGN KEY (batch_id) REFERENCES stock_batches (id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )
    migrate_products_table(db)
    migrate_product_unique_index(db)
    migrate_jobs_table(db)
    migrate_job_materials_table(db)
    migrate_settings_table(db)
    migrate_stock_batches(db)
    db.commit()
    db.close()


def list_backups(limit: int = 8) -> list[Path]:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backups = sorted(BACKUP_DIR.glob("stock_manager_backup_*.db"), reverse=True)
    return backups[:limit]


def get_settings() -> dict[str, str]:
    settings = DEFAULT_SETTINGS.copy()
    rows = get_db().execute("SELECT key, value FROM app_settings").fetchall()
    for row in rows:
        settings[row["key"]] = row["value"]
    return settings


def save_settings(form_data: Any) -> None:
    db = get_db()
    normalized_categories: list[str] = []
    for key, default_value in DEFAULT_SETTINGS.items():
        value = form_data.get(key, "").strip() or default_value
        if key == "product_categories":
            normalized_lines: list[str] = []
            for raw_line in value.splitlines():
                category = raw_line.strip()
                if category and category not in normalized_lines:
                    normalized_lines.append(category)
            value = "\n".join(normalized_lines)
            normalized_categories = normalized_lines
        db.execute(
            """
            INSERT INTO app_settings (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (key, value),
        )
    if normalized_categories:
        placeholders = ",".join("?" for _ in normalized_categories)
        db.execute(
            f"UPDATE products SET category = '' WHERE category <> '' AND category NOT IN ({placeholders})",
            normalized_categories,
        )
    else:
        db.execute("UPDATE products SET category = '' WHERE category <> ''")
    db.commit()


def create_database_backup() -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
    backup_path = BACKUP_DIR / f"stock_manager_backup_{timestamp}.db"
    if "db" in g:
        g.db.commit()
    if DATABASE_PATH.exists():
        shutil.copy2(DATABASE_PATH, backup_path)
    else:
        sqlite3.connect(backup_path).close()
    return backup_path


def get_backup_path_or_404(backup_name: str) -> Path:
    backup_path = (BACKUP_DIR / backup_name).resolve()
    allowed_root = BACKUP_DIR.resolve()
    if allowed_root not in backup_path.parents:
        abort(404)
    if not backup_path.exists() or backup_path.suffix != ".db":
        abort(404)
    return backup_path


def restore_database_backup(backup_name: str) -> Path:
    backup_path = get_backup_path_or_404(backup_name)
    safety_backup = create_database_backup()
    db = g.pop("db", None)
    if db is not None:
        db.commit()
        db.close()
    source_db = sqlite3.connect(backup_path)
    target_db = sqlite3.connect(DATABASE_PATH)
    try:
        source_db.backup(target_db)
    finally:
        target_db.close()
        source_db.close()
    return safety_backup


def reset_database_data() -> None:
    db = get_db()
    db.execute("DELETE FROM job_material_batch_allocations")
    db.execute("DELETE FROM job_materials")
    db.execute("DELETE FROM stock_batches")
    db.execute("DELETE FROM stock_purchases")
    db.execute("DELETE FROM jobs")
    db.execute("DELETE FROM products")
    db.execute("DELETE FROM app_settings")
    db.execute(
        "DELETE FROM sqlite_sequence WHERE name IN ('products', 'jobs', 'job_materials', 'stock_purchases', 'stock_batches', 'job_material_batch_allocations')"
    )
    db.commit()


def migrate_products_table(db: sqlite3.Connection) -> None:
    existing_columns = {
        row[1] for row in db.execute("PRAGMA table_info(products)").fetchall()
    }
    if "barcode" not in existing_columns:
        db.execute("ALTER TABLE products ADD COLUMN barcode TEXT NOT NULL DEFAULT ''")
    if "image_url" not in existing_columns:
        db.execute("ALTER TABLE products ADD COLUMN image_url TEXT NOT NULL DEFAULT ''")
    if "purchase_quantity" not in existing_columns:
        db.execute(
            "ALTER TABLE products ADD COLUMN purchase_quantity REAL NOT NULL DEFAULT 1"
        )
    if "purchase_price" not in existing_columns:
        db.execute(
            "ALTER TABLE products ADD COLUMN purchase_price REAL NOT NULL DEFAULT 0"
        )
        db.execute("UPDATE products SET purchase_price = cost WHERE purchase_price = 0")
    if "stock_quantity" not in existing_columns:
        db.execute(
            "ALTER TABLE products ADD COLUMN stock_quantity REAL NOT NULL DEFAULT 0"
        )
    if "meter_tracking_enabled" not in existing_columns:
        db.execute(
            "ALTER TABLE products ADD COLUMN meter_tracking_enabled INTEGER NOT NULL DEFAULT 0"
        )

    db.execute(
        """
        UPDATE products
        SET
            barcode = CASE
                WHEN barcode IS NULL THEN ''
                ELSE TRIM(barcode)
            END,
            purchase_quantity = CASE
                WHEN purchase_quantity IS NULL OR purchase_quantity <= 0 THEN 1
                ELSE purchase_quantity
            END,
            purchase_price = CASE
                WHEN purchase_price IS NULL OR purchase_price < 0 THEN cost
                ELSE purchase_price
            END,
            stock_quantity = CASE
                WHEN stock_quantity IS NULL THEN 0
                ELSE stock_quantity
            END
        """
    )
    db.execute("UPDATE products SET cost = ROUND(purchase_price / purchase_quantity, 4)")


def migrate_product_unique_index(db: sqlite3.Connection) -> None:
    duplicates = db.execute(
        """
        SELECT LOWER(TRIM(article_number)) AS normalized_article_number, COUNT(*) AS duplicate_count
        FROM products
        GROUP BY LOWER(TRIM(article_number))
        HAVING normalized_article_number <> '' AND duplicate_count > 1
        """
    ).fetchall()
    if duplicates:
        return
    db.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_products_article_number_unique ON products(article_number COLLATE NOCASE)"
    )


def migrate_jobs_table(db: sqlite3.Connection) -> None:
    existing_columns = {row[1] for row in db.execute("PRAGMA table_info(jobs)").fetchall()}
    if "is_archived" not in existing_columns:
        db.execute("ALTER TABLE jobs ADD COLUMN is_archived INTEGER NOT NULL DEFAULT 0")


def migrate_job_materials_table(db: sqlite3.Connection) -> None:
    existing_columns = {
        row[1] for row in db.execute("PRAGMA table_info(job_materials)").fetchall()
    }
    if "is_invoiced" not in existing_columns:
        db.execute("ALTER TABLE job_materials ADD COLUMN is_invoiced INTEGER NOT NULL DEFAULT 0")
    if "invoice_number" not in existing_columns:
        db.execute("ALTER TABLE job_materials ADD COLUMN invoice_number TEXT NOT NULL DEFAULT ''")
    if "meter_start_snapshot" not in existing_columns:
        db.execute("ALTER TABLE job_materials ADD COLUMN meter_start_snapshot REAL")
    if "meter_end_snapshot" not in existing_columns:
        db.execute("ALTER TABLE job_materials ADD COLUMN meter_end_snapshot REAL")


def migrate_stock_batches(db: sqlite3.Connection) -> None:
    products = db.execute(
        """
        SELECT id, stock_quantity, cost, purchase_price, purchase_quantity
        FROM products
        """
    ).fetchall()
    for product in products:
        batch_row = db.execute(
            "SELECT COUNT(*) AS count FROM stock_batches WHERE product_id = ?",
            (product["id"],),
        ).fetchone()
        if batch_row["count"] == 0 and product["stock_quantity"] > 0:
            fallback_cost = product["cost"] or calculate_unit_cost(
                product["purchase_price"], product["purchase_quantity"]
            )
            db.execute(
                """
                INSERT INTO stock_batches (product_id, quantity_remaining, unit_cost, source_type)
                VALUES (?, ?, ?, 'opening')
                """,
                (product["id"], product["stock_quantity"], fallback_cost),
            )


def migrate_settings_table(db: sqlite3.Connection) -> None:
    for key, legacy_value in LEGACY_DEFAULT_SETTINGS.items():
        current_row = db.execute(
            "SELECT value FROM app_settings WHERE key = ?",
            (key,),
        ).fetchone()
        if current_row and current_row[0] == legacy_value:
            db.execute(
                "UPDATE app_settings SET value = ? WHERE key = ?",
                (DEFAULT_SETTINGS[key], key),
            )


def query_categories() -> list[str]:
    settings = get_settings()
    return [
        line.strip()
        for line in settings.get("product_categories", "").splitlines()
        if line.strip()
    ]


def query_recent_purchases(limit: int = 10) -> list[sqlite3.Row]:
    return get_db().execute(
        """
        SELECT
            stock_purchases.*,
            products.description,
            products.article_number
        FROM stock_purchases
        INNER JOIN products ON products.id = stock_purchases.product_id
        ORDER BY stock_purchases.created_at DESC, stock_purchases.id DESC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()


def query_stock_products(search: str = "") -> list[sqlite3.Row]:
    conditions: list[str] = []
    params: list[Any] = []
    if search:
        conditions.append("(article_number LIKE ? OR description LIKE ?)")
        pattern = f"%{search}%"
        params.extend([pattern, pattern])

    where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    return get_db().execute(
        f"""
        SELECT *
        FROM products
        {where_sql}
        ORDER BY category COLLATE NOCASE, description COLLATE NOCASE, article_number COLLATE NOCASE
        """,
        params,
    ).fetchall()


def get_selected_stock_product(product_id: int | None, products: list[sqlite3.Row]) -> sqlite3.Row | None:
    if not products:
        return None
    if product_id is None:
        return products[0]
    for product in products:
        if product["id"] == product_id:
            return product
    return products[0]


def get_product_or_404(product_id: int) -> sqlite3.Row:
    product = get_db().execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if product is None:
        abort(404)
    return product


def get_product_by_article_number(article_number: str, exclude_product_id: int | None = None) -> sqlite3.Row | None:
    normalized_article_number = article_number.strip()
    if not normalized_article_number:
        return None

    params: list[Any] = [normalized_article_number]
    sql = """
        SELECT *
        FROM products
        WHERE LOWER(TRIM(article_number)) = LOWER(TRIM(?))
    """
    if exclude_product_id is not None:
        sql += " AND id <> ?"
        params.append(exclude_product_id)
    sql += " ORDER BY id ASC LIMIT 1"
    return get_db().execute(sql, params).fetchone()


def get_job_or_404(job_id: int) -> sqlite3.Row:
    job = get_db().execute("SELECT * FROM jobs WHERE id = ?", (job_id,)).fetchone()
    if job is None:
        abort(404)
    return job


def calculate_sale_price(cost: float, margin: float) -> float:
    return round(cost * (1 + (margin / 100.0)), 2)


def calculate_unit_cost(purchase_price: float, purchase_quantity: float) -> float:
    normalized_quantity = purchase_quantity if purchase_quantity > 0 else 1
    return round(purchase_price / normalized_quantity, 4)


def create_stock_batch(
    db: sqlite3.Connection,
    product_id: int,
    quantity: float,
    unit_cost: float,
    source_type: str,
) -> None:
    if quantity <= 0:
        return
    db.execute(
        """
        INSERT INTO stock_batches (product_id, quantity_remaining, unit_cost, source_type)
        VALUES (?, ?, ?, ?)
        """,
        (product_id, quantity, unit_cost, source_type),
    )


def get_product_batch_stock(db: sqlite3.Connection, product_id: int) -> float:
    row = db.execute(
        "SELECT COALESCE(SUM(quantity_remaining), 0) AS total FROM stock_batches WHERE product_id = ?",
        (product_id,),
    ).fetchone()
    return float(row["total"])


def get_display_unit_cost(
    db: sqlite3.Connection,
    product_id: int,
    fallback_purchase_price: float,
    fallback_purchase_quantity: float,
) -> float:
    batch = db.execute(
        """
        SELECT unit_cost
        FROM stock_batches
        WHERE product_id = ? AND quantity_remaining > 0
        ORDER BY id ASC
        LIMIT 1
        """,
        (product_id,),
    ).fetchone()
    if batch is not None:
        return round(float(batch["unit_cost"]), 4)
    return calculate_unit_cost(fallback_purchase_price, fallback_purchase_quantity)


def refresh_product_stock_state(db: sqlite3.Connection, product_id: int) -> None:
    product = db.execute(
        "SELECT purchase_price, purchase_quantity FROM products WHERE id = ?",
        (product_id,),
    ).fetchone()
    if product is None:
        return
    stock_quantity = get_product_batch_stock(db, product_id)
    unit_cost = get_display_unit_cost(
        db,
        product_id,
        float(product["purchase_price"]),
        float(product["purchase_quantity"]),
    )
    db.execute(
        "UPDATE products SET stock_quantity = ?, cost = ? WHERE id = ?",
        (stock_quantity, unit_cost, product_id),
    )


def consume_stock_fifo(
    db: sqlite3.Connection,
    product_id: int,
    quantity: float,
) -> tuple[list[tuple[int | None, float, float]], float]:
    remaining = quantity
    allocations: list[tuple[int | None, float, float]] = []
    weighted_cost_total = 0.0
    batches = db.execute(
        """
        SELECT id, quantity_remaining, unit_cost
        FROM stock_batches
        WHERE product_id = ? AND quantity_remaining > 0
        ORDER BY id ASC
        """,
        (product_id,),
    ).fetchall()

    for batch in batches:
        if remaining <= 0:
            break
        take_quantity = min(float(batch["quantity_remaining"]), remaining)
        if take_quantity <= 0:
            continue
        db.execute(
            "UPDATE stock_batches SET quantity_remaining = quantity_remaining - ? WHERE id = ?",
            (take_quantity, batch["id"]),
        )
        allocations.append((batch["id"], take_quantity, float(batch["unit_cost"])))
        weighted_cost_total += take_quantity * float(batch["unit_cost"])
        remaining -= take_quantity

    if remaining > 0:
        raise ValueError("Onvoldoende voorraad in stock batches.")

    weighted_unit_cost = round(weighted_cost_total / quantity, 4) if quantity > 0 else 0.0
    return allocations, weighted_unit_cost


def restore_material_allocations(
    db: sqlite3.Connection,
    material_id: int,
    product_id: int,
    fallback_quantity: float,
    fallback_unit_cost: float,
) -> None:
    allocations = db.execute(
        """
        SELECT batch_id, quantity, unit_cost_snapshot
        FROM job_material_batch_allocations
        WHERE job_material_id = ?
        ORDER BY id ASC
        """,
        (material_id,),
    ).fetchall()
    if allocations:
        for allocation in allocations:
            if allocation["batch_id"] is not None:
                existing_batch = db.execute(
                    "SELECT id FROM stock_batches WHERE id = ?",
                    (allocation["batch_id"],),
                ).fetchone()
                if existing_batch is not None:
                    db.execute(
                        "UPDATE stock_batches SET quantity_remaining = quantity_remaining + ? WHERE id = ?",
                        (allocation["quantity"], allocation["batch_id"]),
                    )
                    continue
            create_stock_batch(
                db,
                product_id,
                float(allocation["quantity"]),
                float(allocation["unit_cost_snapshot"]),
                "restore",
            )
        db.execute(
            "DELETE FROM job_material_batch_allocations WHERE job_material_id = ?",
            (material_id,),
        )
    elif fallback_quantity > 0:
        create_stock_batch(db, product_id, fallback_quantity, fallback_unit_cost, "restore")


def sync_product_stock_level(
    db: sqlite3.Connection,
    product_id: int,
    target_stock_quantity: float,
    unit_cost: float,
) -> None:
    current_stock = get_product_batch_stock(db, product_id)
    difference = round(target_stock_quantity - current_stock, 4)
    if difference > 0:
        create_stock_batch(db, product_id, difference, unit_cost, "adjustment")
    elif difference < 0:
        consume_stock_fifo(db, product_id, abs(difference))
    refresh_product_stock_state(db, product_id)


def format_quantity(quantity: float) -> str:
    return f"{quantity:g}"


def calculate_stock_value(unit_cost: float, stock_quantity: float) -> float:
    return round(unit_cost * stock_quantity, 2)


def parse_decimal(value: str, field_name: str) -> float:
    normalized = (value or "").strip().replace(",", ".")
    if normalized == "":
        raise ValueError(f"Ontbrekende waarde voor {field_name}.")
    try:
        return float(normalized)
    except ValueError as exc:
        raise ValueError(f"Ongeldig getal voor {field_name}: {value}") from exc


def parse_product_form(form: Any) -> dict[str, Any]:
    article_number = form.get("article_number", "").strip()
    description = form.get("description", "").strip()
    unit = form.get("unit", "").strip()
    if not article_number:
        raise ValueError("Artikelnummer is verplicht.")
    if not description:
        raise ValueError("Omschrijving is verplicht.")
    if not unit:
        raise ValueError("Verkoopeenheid is verplicht.")

    purchase_quantity = float(form.get("purchase_quantity", 1) or 1)
    purchase_price = float(form.get("purchase_price", 0) or 0)
    stock_quantity = float(form.get("stock_quantity", 0) or 0)
    profit_margin = float(form.get("profit_margin", 0) or 0)
    if purchase_quantity <= 0:
        raise ValueError("Aantal per aankoop moet groter zijn dan nul.")
    if purchase_price < 0:
        raise ValueError("Aankoopprijs mag niet negatief zijn.")
    if stock_quantity < 0:
        raise ValueError("Voorraad mag niet negatief zijn.")
    if profit_margin < 0:
        raise ValueError("Winstmarge mag niet negatief zijn.")

    return {
        "article_number": article_number,
        "barcode": form.get("barcode", "").strip(),
        "image_url": form.get("image_url", "").strip(),
        "description": description,
        "unit": unit,
        "purchase_quantity": purchase_quantity,
        "purchase_price": purchase_price,
        "stock_quantity": stock_quantity,
        "cost": calculate_unit_cost(purchase_price, purchase_quantity),
        "profit_margin": profit_margin,
        "meter_tracking_enabled": 1 if form.get("meter_tracking_enabled") == "on" else 0,
        "category": normalize_category_value(form.get("category", "")),
    }


def save_product_record(product_data: dict[str, Any], product_id: int | None = None) -> int:
    db = get_db()
    if product_id is None:
        cursor = db.execute(
            """
            INSERT INTO products (
                article_number,
                barcode,
                image_url,
                description,
                unit,
                purchase_quantity,
                purchase_price,
                stock_quantity,
                cost,
                profit_margin,
                meter_tracking_enabled,
                category
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                product_data["article_number"],
                product_data["barcode"],
                product_data["image_url"],
                product_data["description"],
                product_data["unit"],
                product_data["purchase_quantity"],
                product_data["purchase_price"],
                product_data["stock_quantity"],
                product_data["cost"],
                product_data["profit_margin"],
                product_data["meter_tracking_enabled"],
                product_data["category"],
            ),
        )
        new_product_id = int(cursor.lastrowid)
        if product_data["stock_quantity"] > 0:
            create_stock_batch(db, new_product_id, product_data["stock_quantity"], product_data["cost"], "opening")
            refresh_product_stock_state(db, new_product_id)
        return new_product_id

    db.execute(
        """
        UPDATE products
        SET article_number = ?, barcode = ?, image_url = ?, description = ?, unit = ?, purchase_quantity = ?, purchase_price = ?,
            stock_quantity = ?, cost = ?, profit_margin = ?, meter_tracking_enabled = ?, category = ?
        WHERE id = ?
        """,
        (
            product_data["article_number"],
            product_data["barcode"],
            product_data["image_url"],
            product_data["description"],
            product_data["unit"],
            product_data["purchase_quantity"],
            product_data["purchase_price"],
            product_data["stock_quantity"],
            product_data["cost"],
            product_data["profit_margin"],
            product_data["meter_tracking_enabled"],
            product_data["category"],
            product_id,
        ),
    )
    sync_product_stock_level(db, product_id, product_data["stock_quantity"], product_data["cost"])
    return product_id


def normalize_product_payload(raw_row: dict[str, str]) -> dict[str, Any]:
    row = {key.strip(): (value or "").strip() for key, value in raw_row.items() if key}
    required_text_fields = ["article_number", "description", "unit"]
    for field_name in required_text_fields:
        if not row.get(field_name):
            raise ValueError(f"Ontbrekende waarde voor {field_name}.")

    purchase_quantity = parse_decimal(row.get("purchase_quantity", ""), "purchase_quantity")
    if purchase_quantity <= 0:
        raise ValueError("purchase_quantity moet groter zijn dan nul.")

    purchase_price = parse_decimal(row.get("purchase_price", ""), "purchase_price")
    if purchase_price < 0:
        raise ValueError("purchase_price mag niet negatief zijn.")

    stock_quantity = parse_decimal(row.get("stock_quantity", "0"), "stock_quantity")
    if stock_quantity < 0:
        raise ValueError("stock_quantity mag niet negatief zijn.")

    profit_margin = parse_decimal(row.get("profit_margin", "0"), "profit_margin")
    if profit_margin < 0:
        raise ValueError("profit_margin mag niet negatief zijn.")

    meter_tracking_enabled = row.get("meter_tracking_enabled", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
    }

    return {
        "article_number": row["article_number"],
        "barcode": row.get("barcode", "").strip(),
        "image_url": row.get("image_url", "").strip(),
        "description": row["description"],
        "unit": row["unit"],
        "purchase_quantity": purchase_quantity,
        "purchase_price": purchase_price,
        "stock_quantity": stock_quantity,
        "cost": calculate_unit_cost(purchase_price, purchase_quantity),
        "profit_margin": profit_margin,
        "meter_tracking_enabled": 1 if meter_tracking_enabled else 0,
        "category": normalize_category_value(row.get("category", "")),
    }


def normalize_unit_label(unit: str) -> str:
    normalized = unit.strip().lower()
    unit_map = {
        "st": "st",
        "stuk": "st",
        "stukken": "st",
        "m": "m",
        "meter": "m",
        "meters": "m",
    }
    return unit_map.get(normalized, normalized or "st")


def parse_spreadsheet_decimal(value: Any, field_name: str) -> float:
    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value or "").strip()
    if not normalized:
        raise ValueError(f"Ontbrekende waarde voor {field_name}.")

    normalized = normalized.replace("\xa0", "").replace("€", "").replace(" ", "")
    if re.fullmatch(r"-?\d{1,3}(,\d{3})+", normalized):
        normalized = normalized.replace(",", "")
    elif "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    else:
        normalized = normalized.replace(",", ".")
    normalized = normalized.replace("-.", "-0.")

    try:
        return float(normalized)
    except ValueError as exc:
        raise ValueError(f"Ongeldig getal voor {field_name}: {value}") from exc


def get_default_import_category() -> str:
    return ""


def normalize_category_value(category: str) -> str:
    normalized_category = category.strip()
    if not normalized_category:
        return ""
    return normalized_category if normalized_category in query_categories() else ""


def find_excel_header_row(rows: list[list[Any]], required_headers: set[str]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        header_map: dict[str, int] = {}
        for column_index, cell in enumerate(row):
            header = str(cell or "").strip().lower()
            if header:
                header_map[header] = column_index
        if required_headers.issubset(header_map.keys()):
            return row_index, header_map
    raise ValueError(
        "De Excel-import verwacht kolommen met minstens: Artikel, Omschrijving, Netto Prijs, Per, Aantal en Totaal."
    )


def load_excel_rows(raw_bytes: bytes, filename: str) -> list[list[Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore[import-not-found]
        except ImportError as exc:
            raise ValueError("Excel-import is nog niet beschikbaar omdat het pakket xlrd ontbreekt.") from exc

        workbook = xlrd.open_workbook(file_contents=raw_bytes)
        if workbook.nsheets == 0:
            raise ValueError("Het gekozen Excel-bestand bevat geen werkblad.")
        sheet = workbook.sheet_by_index(0)
        return [sheet.row_values(row_index) for row_index in range(sheet.nrows)]

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:
            raise ValueError("Excel-import is nog niet beschikbaar omdat het pakket openpyxl ontbreekt.") from exc

        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        sheet = workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    raise ValueError("Excel-import ondersteunt enkel .xls en .xlsx bestanden.")


def load_supplier_excel_products(file_storage: Any) -> list[dict[str, Any]]:
    if not file_storage or not file_storage.filename:
        raise ValueError("Kies eerst een Excel-bestand.")

    raw_bytes = file_storage.stream.read()
    if not raw_bytes:
        raise ValueError("Het gekozen Excel-bestand is leeg.")

    rows = load_excel_rows(raw_bytes, file_storage.filename)
    required_headers = {"artikel", "omschrijving", "netto prijs", "per", "aantal", "totaal"}
    header_row_index, header_map = find_excel_header_row(rows, required_headers)
    default_category = get_default_import_category()
    products: list[dict[str, Any]] = []

    for line_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        article_number = str(row[header_map["artikel"]] or "").strip()
        description = str(row[header_map["omschrijving"]] or "").strip()
        if not article_number and not description:
            continue
        if not article_number:
            continue
        if not description:
            raise ValueError(f"Rij {line_number}: ontbrekende omschrijving voor artikel {article_number}.")

        purchase_quantity = parse_spreadsheet_decimal(row[header_map["aantal"]], "aantal")
        if purchase_quantity <= 0:
            raise ValueError(f"Rij {line_number}: aantal moet groter zijn dan nul.")

        total_purchase_price = parse_spreadsheet_decimal(row[header_map["totaal"]], "totaal")
        if total_purchase_price <= 0:
            continue

        unit_raw = str(row[header_map["per"]] or "").strip()
        unit = normalize_unit_label(unit_raw)

        products.append(
            {
                "article_number": article_number,
                "description": description,
                "unit": unit,
                "purchase_quantity": purchase_quantity,
                "purchase_price": total_purchase_price,
                "cost": calculate_unit_cost(total_purchase_price, purchase_quantity),
                "category": default_category,
            }
        )

    if not products:
        raise ValueError("Geen bruikbare productregels gevonden in het Excel-bestand.")
    return products


def import_supplier_excel(file_storage: Any) -> tuple[int, int]:
    products = load_supplier_excel_products(file_storage)
    db = get_db()
    inserted_count = 0
    updated_count = 0

    for product in products:
        existing = db.execute(
            "SELECT id, profit_margin, meter_tracking_enabled, category, barcode FROM products WHERE article_number = ?",
            (product["article_number"],),
        ).fetchone()

        if existing:
            db.execute(
                """
                UPDATE products
                SET description = ?, unit = ?, purchase_quantity = ?, purchase_price = ?
                WHERE id = ?
                """,
                (
                    product["description"],
                    product["unit"],
                    product["purchase_quantity"],
                    product["purchase_price"],
                    existing["id"],
                ),
            )
            refresh_product_stock_state(db, int(existing["id"]))
            updated_count += 1
        else:
            db.execute(
                """
                INSERT INTO products (
                    article_number,
                    barcode,
                    description,
                    unit,
                    purchase_quantity,
                    purchase_price,
                    stock_quantity,
                    cost,
                    profit_margin,
                    meter_tracking_enabled,
                    category
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product["article_number"],
                    "",
                    product["description"],
                    product["unit"],
                    product["purchase_quantity"],
                    product["purchase_price"],
                    0,
                    product["cost"],
                    0,
                    0,
                    product["category"],
                ),
            )
            inserted_count += 1

    db.commit()
    return inserted_count, updated_count


def import_products_from_csv(file_storage: Any) -> tuple[int, int]:
    if not file_storage or not file_storage.filename:
        raise ValueError("Kies eerst een CSV-bestand.")

    raw_bytes = file_storage.stream.read()
    if not raw_bytes:
        raise ValueError("Het gekozen CSV-bestand is leeg.")

    text = raw_bytes.decode("utf-8-sig")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    expected_headers = {
        "article_number",
        "description",
        "unit",
        "purchase_quantity",
        "purchase_price",
        "stock_quantity",
        "profit_margin",
        "meter_tracking_enabled",
        "category",
    }
    headers = {header.strip() for header in (reader.fieldnames or []) if header}
    missing_headers = sorted(expected_headers - headers)
    if missing_headers:
        raise ValueError(f"Ontbrekende CSV-kolommen: {', '.join(missing_headers)}")
    should_update_image_url = "image_url" in headers

    db = get_db()
    inserted_count = 0
    updated_count = 0

    for line_number, raw_row in enumerate(reader, start=2):
        if raw_row is None:
            continue
        if not any((value or "").strip() for value in raw_row.values()):
            continue
        try:
            product = normalize_product_payload(raw_row)
        except ValueError as exc:
            raise ValueError(f"Line {line_number}: {exc}") from exc

        existing = db.execute(
            "SELECT id, image_url FROM products WHERE article_number = ?",
            (product["article_number"],),
        ).fetchone()

        if existing:
            image_url = product["image_url"] if should_update_image_url else existing["image_url"]
            db.execute(
                """
                UPDATE products
                SET barcode = ?, image_url = ?, description = ?, unit = ?, purchase_quantity = ?, purchase_price = ?,
                    profit_margin = ?, meter_tracking_enabled = ?, category = ?
                WHERE id = ?
                """,
                (
                    product["barcode"],
                    image_url,
                    product["description"],
                    product["unit"],
                    product["purchase_quantity"],
                    product["purchase_price"],
                    product["profit_margin"],
                    product["meter_tracking_enabled"],
                    product["category"],
                    existing["id"],
                ),
            )
            sync_product_stock_level(
                db,
                int(existing["id"]),
                float(product["stock_quantity"]),
                float(product["cost"]),
            )
            updated_count += 1
        else:
            cursor = db.execute(
                """
                INSERT INTO products (
                    article_number,
                    barcode,
                    image_url,
                    description,
                    unit,
                    purchase_quantity,
                    purchase_price,
                    stock_quantity,
                    cost,
                    profit_margin,
                    meter_tracking_enabled,
                    category
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product["article_number"],
                    product["barcode"],
                    product["image_url"],
                    product["description"],
                    product["unit"],
                    product["purchase_quantity"],
                    product["purchase_price"],
                    product["stock_quantity"],
                    product["cost"],
                    product["profit_margin"],
                    product["meter_tracking_enabled"],
                    product["category"],
                ),
            )
            product_id = int(cursor.lastrowid)
            if float(product["stock_quantity"]) > 0:
                create_stock_batch(
                    db,
                    product_id,
                    float(product["stock_quantity"]),
                    float(product["cost"]),
                    "opening",
                )
                refresh_product_stock_state(db, product_id)
            inserted_count += 1

    db.commit()
    return inserted_count, updated_count


@app.context_processor
def inject_helpers() -> dict[str, Any]:
    return {
        "calculate_sale_price": calculate_sale_price,
        "calculate_unit_cost": calculate_unit_cost,
        "calculate_stock_value": calculate_stock_value,
        "format_quantity": format_quantity,
        "database_path": DATABASE_PATH,
        "ui_settings": get_settings(),
        "auth_enabled": is_auth_enabled(),
        "is_logged_in": is_logged_in(),
        "admin_username": ADMIN_USERNAME,
    }


@app.route("/login", methods=["GET", "POST"])
def login() -> str | Response:
    if not is_auth_enabled():
        return redirect(url_for("index"))

    next_url = request.args.get("next", "").strip() or request.form.get("next", "").strip()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["authenticated"] = True
            session["username"] = username
            flash("Aangemeld.", "success")
            if next_url.startswith("/"):
                return redirect(next_url)
            return redirect(url_for("index"))
        flash("Ongeldige gebruikersnaam of wachtwoord.", "error")

    return render_template("login.html", next_url=next_url)


@app.post("/logout")
def logout() -> Response:
    session.clear()
    flash("Afgemeld.", "success")
    return redirect(url_for("login"))


@app.route("/")
def index() -> str:
    db = get_db()
    total_products = db.execute("SELECT COUNT(*) AS count FROM products").fetchone()["count"]
    total_jobs = db.execute("SELECT COUNT(*) AS count FROM jobs").fetchone()["count"]
    stock_value = db.execute(
        "SELECT COALESCE(SUM(quantity_remaining * unit_cost), 0) AS total FROM stock_batches"
    ).fetchone()["total"]
    low_stock_count = db.execute(
        "SELECT COUNT(*) AS count FROM products WHERE stock_quantity <= 0"
    ).fetchone()["count"]
    return render_template(
        "index.html",
        total_products=total_products,
        total_jobs=total_jobs,
        stock_value=round(stock_value, 2),
        low_stock_count=low_stock_count,
    )


@app.route("/products")
def products() -> str:
    sort = request.args.get("sort", "description")
    direction = request.args.get("direction", "asc")
    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()

    allowed_sort_columns = {
        "article_number": "article_number",
        "description": "description",
        "unit": "unit",
        "purchase_quantity": "purchase_quantity",
        "purchase_price": "purchase_price",
        "stock_quantity": "stock_quantity",
        "cost": "cost",
        "profit_margin": "profit_margin",
        "category": "category",
        "created_at": "created_at",
    }
    sort_column = allowed_sort_columns.get(sort, "description")
    sort_direction = "DESC" if direction.lower() == "desc" else "ASC"

    conditions: list[str] = []
    params: list[Any] = []
    if search:
        conditions.append("(article_number LIKE ? OR barcode LIKE ? OR description LIKE ? OR category LIKE ?)")
        pattern = f"%{search}%"
        params.extend([pattern, pattern, pattern, pattern])
    if category:
        conditions.append("category = ?")
        params.append(category)

    where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    sql = f"""
        SELECT *
        FROM products
        {where_sql}
        ORDER BY
            CASE WHEN TRIM(category) = '' THEN 1 ELSE 0 END ASC,
            category COLLATE NOCASE ASC,
            {sort_column} {sort_direction},
            id DESC
    """
    rows = get_db().execute(sql, params).fetchall()
    grouped_products: list[tuple[str, list[sqlite3.Row]]] = []
    current_group_label = None
    current_group_rows: list[sqlite3.Row] = []
    for row in rows:
        group_label = row["category"].strip() or "Zonder categorie"
        if current_group_label != group_label:
            if current_group_rows:
                grouped_products.append((current_group_label, current_group_rows))
            current_group_label = group_label
            current_group_rows = [row]
        else:
            current_group_rows.append(row)
    if current_group_rows:
        grouped_products.append((current_group_label, current_group_rows))

    return render_template(
        "products.html",
        products=rows,
        grouped_products=grouped_products,
        categories=query_categories(),
        selected_sort=sort,
        selected_direction=direction,
        selected_search=search,
        selected_category=category,
    )


@app.route("/stock")
def stock() -> Response:
    return redirect(url_for("products"))


@app.post("/backups")
def backup_database() -> Response:
    backup_path = create_database_backup()
    flash(f"Back-up aangemaakt: {backup_path.name}", "success")
    return redirect(url_for("settings"))


@app.post("/backups/<path:backup_name>/restore")
def restore_backup_database(backup_name: str) -> Response:
    safety_backup = restore_database_backup(backup_name)
    flash(
        f"Back-up teruggezet: {backup_name}. Voor de zekerheid werd eerst een extra back-up gemaakt: {safety_backup.name}.",
        "success",
    )
    return redirect(url_for("settings"))


@app.route("/settings", methods=["GET", "POST"])
def settings() -> str | Response:
    if request.method == "POST":
        save_settings(request.form)
        flash("Instellingen opgeslagen.", "success")
        return redirect(url_for("settings"))
    return render_template(
        "settings.html",
        settings=get_settings(),
        backup_files=list_backups(),
    )


@app.post("/settings/reset-database")
def reset_database() -> Response:
    reset_database_data()
    flash("Alle gegevens zijn verwijderd. De app is teruggezet naar een lege standaarddatabase.", "success")
    return redirect(url_for("settings"))


@app.post("/products")
def create_product() -> Response:
    try:
        product_data = parse_product_form(request.form)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("products"))

    duplicate_product = get_product_by_article_number(product_data["article_number"])
    if duplicate_product is not None and request.form.get("confirm_overwrite") != "1":
        flash("Dit artikelnummer bestaat al. Kies overschrijven of annuleer.", "error")
        return redirect(url_for("products"))

    save_product_record(product_data, int(duplicate_product["id"])) if duplicate_product is not None else save_product_record(product_data)
    get_db().commit()
    if duplicate_product is not None:
        flash("Bestaand product overschreven.", "success")
    else:
        flash("Product toegevoegd.", "success")
    return redirect(url_for("products"))


@app.post("/products/import")
def import_products() -> Response:
    import_file = request.files.get("csv_file")
    filename = (import_file.filename or "").lower() if import_file else ""
    try:
        if filename.endswith(".csv"):
            inserted_count, updated_count = import_products_from_csv(import_file)
            success_message = (
                f"CSV-import voltooid. {inserted_count} producten toegevoegd en {updated_count} producten bijgewerkt."
            )
        elif filename.endswith(".xls") or filename.endswith(".xlsx"):
            inserted_count, updated_count = import_supplier_excel(import_file)
            success_message = (
                f"Excel-import voltooid. {inserted_count} producten toegevoegd en {updated_count} producten bijgewerkt."
            )
        else:
            raise ValueError("Kies een .csv, .xls of .xlsx bestand.")
    except UnicodeDecodeError:
        flash("CSV-import mislukt: bestand moet UTF-8 gecodeerd zijn.", "error")
        return redirect(url_for("products"))
    except ValueError as exc:
        flash(f"Import mislukt: {exc}", "error")
        return redirect(url_for("products"))

    flash(success_message, "success")
    return redirect(url_for("products"))


@app.post("/products/<int:product_id>/update")
def update_product(product_id: int) -> Response:
    get_product_or_404(product_id)
    try:
        product_data = parse_product_form(request.form)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("products"))

    duplicate_product = get_product_by_article_number(product_data["article_number"], exclude_product_id=product_id)
    if duplicate_product is not None:
        flash("Dit artikelnummer bestaat al bij een ander product. Bewerk dat product of gebruik een ander artikelnummer.", "error")
        return redirect(url_for("products"))

    save_product_record(product_data, product_id)
    get_db().commit()
    flash("Product bijgewerkt.", "success")
    return redirect(url_for("products"))


@app.post("/products/<int:product_id>/adjust-stock")
def adjust_stock(product_id: int) -> Response:
    product = get_product_or_404(product_id)
    adjustment = float(request.form.get("adjustment_quantity", 0) or 0)
    next_page = request.form.get("next_page", "products").strip() or "products"
    if adjustment < 0 and abs(adjustment) > product["stock_quantity"]:
        flash("Voorraadcorrectie zou de voorraad negatief maken. Wijziging niet opgeslagen.", "error")
        return redirect(url_for("products"))

    db = get_db()
    if adjustment > 0:
        create_stock_batch(db, product_id, adjustment, float(product["cost"]), "adjustment")
    elif adjustment < 0:
        consume_stock_fifo(db, product_id, abs(adjustment))
    refresh_product_stock_state(db, product_id)
    db.commit()
    flash("Voorraad bijgewerkt.", "success")
    return redirect(url_for("products"))


@app.post("/products/<int:product_id>/receive-stock")
def receive_stock(product_id: int) -> Response:
    product = get_product_or_404(product_id)
    packages_received = float(request.form.get("packages_received", 0) or 0)
    notes = request.form.get("purchase_notes", "").strip()
    next_page = request.form.get("next_page", "products").strip() or "products"
    purchase_price_raw = request.form.get("purchase_price", "").strip()
    if packages_received <= 0:
        flash("Het aantal ontvangen verpakkingen moet groter zijn dan nul.", "error")
        return redirect(url_for("products"))

    if purchase_price_raw:
        try:
            purchase_price = parse_decimal(purchase_price_raw, "purchase price")
        except ValueError:
            flash("De aankoopprijs moet een geldig getal zijn.", "error")
            return redirect(url_for("products"))
        if purchase_price < 0:
            flash("De aankoopprijs mag niet negatief zijn.", "error")
            return redirect(url_for("products"))
    else:
        purchase_price = float(product["purchase_price"])

    units_added = round(packages_received * product["purchase_quantity"], 4)
    unit_cost = calculate_unit_cost(purchase_price, product["purchase_quantity"])
    db = get_db()
    db.execute(
        """
        INSERT INTO stock_purchases (
            product_id,
            packages_received,
            units_added,
            package_price_snapshot,
            unit_snapshot,
            notes
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            product_id,
            packages_received,
            units_added,
            purchase_price,
            product["unit"],
            notes,
        ),
    )
    db.execute(
        "UPDATE products SET purchase_price = ? WHERE id = ?",
        (purchase_price, product_id),
    )
    create_stock_batch(db, product_id, units_added, unit_cost, "purchase")
    refresh_product_stock_state(db, product_id)
    db.commit()
    flash(
        f"{format_quantity(packages_received)} aankoopseenheden ontvangen en {format_quantity(units_added)} {product['unit']} aan voorraad toegevoegd.",
        "success",
    )
    return redirect(url_for("products"))


@app.post("/products/<int:product_id>/delete")
def delete_product(product_id: int) -> Response:
    db = get_db()
    linked_row = db.execute(
        "SELECT COUNT(*) AS count FROM job_materials WHERE product_id = ?",
        (product_id,),
    ).fetchone()
    if linked_row["count"] == 0:
        db.execute("DELETE FROM products WHERE id = ?", (product_id,))
        db.commit()
    return redirect(url_for("products"))


@app.route("/jobs")
def jobs() -> str:
    view = request.args.get("view", "active")
    archive_filter = "WHERE jobs.is_archived = 1" if view == "archived" else "WHERE jobs.is_archived = 0"
    rows = get_db().execute(
        f"""
        SELECT
            jobs.*,
            COUNT(job_materials.id) AS line_count,
            COALESCE(SUM(job_materials.quantity * job_materials.unit_cost_snapshot), 0) AS total_cost,
            COALESCE(
                SUM(
                    ROUND(
                        job_materials.quantity
                        * ROUND(
                            job_materials.unit_cost_snapshot
                            * (1 + (job_materials.profit_margin_snapshot / 100.0)),
                            2
                        ),
                        2
                    )
                ),
                0
            ) AS total_sale,
            COALESCE(
                SUM(
                    ROUND(
                        job_materials.quantity
                        * ROUND(
                            job_materials.unit_cost_snapshot
                            * (1 + (job_materials.profit_margin_snapshot / 100.0)),
                            2
                        ),
                        2
                    ) - ROUND(job_materials.quantity * job_materials.unit_cost_snapshot, 2)
                ),
                0
            ) AS total_profit
        FROM jobs
        LEFT JOIN job_materials ON job_materials.job_id = jobs.id
        {archive_filter}
        GROUP BY jobs.id
        ORDER BY jobs.created_at DESC, jobs.id DESC
        """
    ).fetchall()
    return render_template("jobs.html", jobs=rows, current_view=view)


@app.post("/jobs")
def create_job() -> Response:
    form = request.form
    db = get_db()
    db.execute(
        "INSERT INTO jobs (name, client_name, notes) VALUES (?, ?, ?)",
        (
            form.get("name", "").strip(),
            form.get("client_name", "").strip(),
            form.get("notes", "").strip(),
        ),
    )
    db.commit()
    return redirect(url_for("jobs"))


@app.post("/jobs/<int:job_id>/update")
def update_job(job_id: int) -> Response:
    get_job_or_404(job_id)
    form = request.form
    db = get_db()
    db.execute(
        """
        UPDATE jobs
        SET name = ?, client_name = ?, notes = ?
        WHERE id = ?
        """,
        (
            form.get("name", "").strip(),
            form.get("client_name", "").strip(),
            form.get("notes", "").strip(),
            job_id,
        ),
    )
    db.commit()
    flash("Werf bijgewerkt.", "success")
    return redirect(url_for("job_detail", job_id=job_id))


@app.post("/jobs/<int:job_id>/archive")
def archive_job(job_id: int) -> Response:
    job = get_job_or_404(job_id)
    new_value = 0 if job["is_archived"] else 1
    db = get_db()
    db.execute("UPDATE jobs SET is_archived = ? WHERE id = ?", (new_value, job_id))
    db.commit()
    flash("Werf gearchiveerd." if new_value else "Werf opnieuw actief gemaakt.", "success")
    return redirect(url_for("jobs", view="archived" if new_value else "active"))


@app.post("/jobs/<int:job_id>/delete")
def delete_job(job_id: int) -> Response:
    get_job_or_404(job_id)
    db = get_db()
    materials = db.execute(
        "SELECT id, product_id, quantity, unit_cost_snapshot FROM job_materials WHERE job_id = ?",
        (job_id,),
    ).fetchall()
    for material in materials:
        restore_material_allocations(
            db,
            int(material["id"]),
            int(material["product_id"]),
            float(material["quantity"]),
            float(material["unit_cost_snapshot"]),
        )
        refresh_product_stock_state(db, int(material["product_id"]))
    db.execute("DELETE FROM jobs WHERE id = ?", (job_id,))
    db.commit()
    flash("Werf verwijderd en voorraad hersteld.", "success")
    return redirect(url_for("jobs"))


@app.route("/jobs/<int:job_id>")
def job_detail(job_id: int) -> str:
    job = get_job_or_404(job_id)
    db = get_db()
    materials = db.execute(
        """
        SELECT
            job_materials.*,
            products.meter_tracking_enabled AS product_meter_tracking_enabled
        FROM job_materials
        LEFT JOIN products ON products.id = job_materials.product_id
        WHERE job_id = ?
        ORDER BY id DESC
        """,
        (job_id,),
    ).fetchall()
    products = db.execute(
        """
        SELECT
            products.*,
            (
                SELECT job_materials.meter_end_snapshot
                FROM job_materials
                WHERE job_materials.product_id = products.id
                    AND job_materials.meter_end_snapshot IS NOT NULL
                ORDER BY job_materials.id DESC
                LIMIT 1
            ) AS last_meter_end
        FROM products
        ORDER BY stock_quantity ASC, description COLLATE NOCASE, article_number COLLATE NOCASE
        """
    ).fetchall()
    totals = db.execute(
        """
        SELECT
            COALESCE(SUM(quantity * unit_cost_snapshot), 0) AS total_cost,
            COALESCE(
                SUM(
                    ROUND(
                        quantity * ROUND(unit_cost_snapshot * (1 + (profit_margin_snapshot / 100.0)), 2),
                        2
                    )
                ),
                0
            ) AS total_sale,
            COALESCE(SUM(quantity), 0) AS total_quantity,
            COALESCE(
                SUM(
                    CASE
                        WHEN is_invoiced = 0 THEN ROUND(
                            quantity * ROUND(unit_cost_snapshot * (1 + (profit_margin_snapshot / 100.0)), 2),
                            2
                        )
                        ELSE 0
                    END
                ),
                0
            ) AS uninvoiced_sale
        FROM job_materials
        WHERE job_id = ?
        """,
        (job_id,),
    ).fetchone()
    return render_template(
        "job_detail.html",
        job=job,
        materials=materials,
        products=products,
        totals=totals,
    )


@app.post("/jobs/<int:job_id>/materials")
def add_job_material(job_id: int) -> Response:
    get_job_or_404(job_id)
    product_id = int(request.form.get("product_id", "0"))
    product = get_product_or_404(product_id)
    quantity_raw = request.form.get("quantity", "").strip()
    meter_start_raw = request.form.get("meter_start", "").strip()
    meter_end_raw = request.form.get("meter_end", "").strip()
    meter_start = None
    meter_end = None

    if meter_start_raw or meter_end_raw:
        if not meter_start_raw or not meter_end_raw:
            flash("Vul beide kabelnummers in of laat ze leeg en geef handmatig het aantal meters op.", "error")
            return redirect(url_for("job_detail", job_id=job_id))
        try:
            meter_start = float(meter_start_raw.replace(",", "."))
            meter_end = float(meter_end_raw.replace(",", "."))
        except ValueError:
            flash("Kabelnummers moeten geldige getallen zijn.", "error")
            return redirect(url_for("job_detail", job_id=job_id))
        quantity = meter_end - meter_start
    else:
        try:
            quantity = parse_decimal(quantity_raw, "quantity used")
        except ValueError:
            flash("Het gebruikte aantal moet een geldig getal zijn.", "error")
            return redirect(url_for("job_detail", job_id=job_id))

    if quantity <= 0:
        if meter_start is not None and meter_end is not None:
            flash("Het eindnummer van de kabel moet groter zijn dan het startnummer.", "error")
        else:
            flash("Het gebruikte aantal moet groter zijn dan nul.", "error")
        return redirect(url_for("job_detail", job_id=job_id))
    if quantity > product["stock_quantity"]:
        flash(
            f"Onvoldoende voorraad voor {product['description']}. Beschikbaar: {format_quantity(product['stock_quantity'])} {product['unit']}.",
            "error",
        )
        return redirect(url_for("job_detail", job_id=job_id))

    db = get_db()
    try:
        allocations, weighted_unit_cost = consume_stock_fifo(db, product_id, quantity)
    except ValueError:
        flash(
            f"Onvoldoende voorraad voor {product['description']}. Beschikbaar: {format_quantity(product['stock_quantity'])} {product['unit']}.",
            "error",
        )
        return redirect(url_for("job_detail", job_id=job_id))

    cursor = db.execute(
        """
        INSERT INTO job_materials (
            job_id,
            product_id,
            quantity,
            article_number_snapshot,
            description_snapshot,
            unit_snapshot,
            unit_cost_snapshot,
            profit_margin_snapshot,
            meter_start_snapshot,
            meter_end_snapshot
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            job_id,
            product_id,
            quantity,
            product["article_number"],
            product["description"],
            product["unit"],
            weighted_unit_cost,
            product["profit_margin"],
            meter_start,
            meter_end,
        ),
    )
    material_id = int(cursor.lastrowid)
    for batch_id, allocation_quantity, allocation_unit_cost in allocations:
        db.execute(
            """
            INSERT INTO job_material_batch_allocations (
                job_material_id,
                batch_id,
                quantity,
                unit_cost_snapshot
            )
            VALUES (?, ?, ?, ?)
            """,
            (material_id, batch_id, allocation_quantity, allocation_unit_cost),
        )
    refresh_product_stock_state(db, product_id)
    db.commit()
    flash("Materiaal toegevoegd en voorraad bijgewerkt.", "success")
    return redirect(url_for("job_detail", job_id=job_id))


@app.post("/jobs/<int:job_id>/materials/<int:material_id>/invoice")
def toggle_job_material_invoiced(job_id: int, material_id: int) -> Response:
    get_job_or_404(job_id)
    db = get_db()
    material = db.execute(
        "SELECT is_invoiced, invoice_number FROM job_materials WHERE id = ? AND job_id = ?",
        (material_id, job_id),
    ).fetchone()
    if material is None:
        abort(404)
    if material["is_invoiced"]:
        db.execute(
            "UPDATE job_materials SET is_invoiced = 0, invoice_number = '' WHERE id = ? AND job_id = ?",
            (material_id, job_id),
        )
        flash("Materiaal gemarkeerd als niet gefactureerd.", "success")
    else:
        invoice_number = request.form.get("invoice_number", "").strip()
        if not invoice_number:
            flash("Vul een factuurnummer in.", "error")
            return redirect(url_for("job_detail", job_id=job_id))
        db.execute(
            "UPDATE job_materials SET is_invoiced = 1, invoice_number = ? WHERE id = ? AND job_id = ?",
            (invoice_number, material_id, job_id),
        )
        flash("Materiaal gemarkeerd als gefactureerd.", "success")
    db.commit()
    return redirect(url_for("job_detail", job_id=job_id))


@app.post("/jobs/<int:job_id>/materials/bulk-invoice")
def bulk_invoice_job_materials(job_id: int) -> Response:
    get_job_or_404(job_id)
    invoice_number = request.form.get("invoice_number", "").strip()
    selected_material_ids: list[int] = []

    for raw_id in request.form.getlist("material_ids"):
        try:
            selected_material_ids.append(int(raw_id))
        except ValueError:
            continue

    if not invoice_number:
        flash("Vul een factuurnummer in.", "error")
        return redirect(url_for("job_detail", job_id=job_id))

    if not selected_material_ids:
        flash("Selecteer minstens een materiaalregel om aan de factuur te koppelen.", "error")
        return redirect(url_for("job_detail", job_id=job_id))

    placeholders = ",".join("?" for _ in selected_material_ids)
    db = get_db()
    matching_materials = db.execute(
        f"""
        SELECT id
        FROM job_materials
        WHERE job_id = ?
            AND is_invoiced = 0
            AND id IN ({placeholders})
        """,
        [job_id, *selected_material_ids],
    ).fetchall()

    if not matching_materials:
        flash("Geen geldige niet-gefactureerde materiaalregels geselecteerd.", "error")
        return redirect(url_for("job_detail", job_id=job_id))

    valid_material_ids = [row["id"] for row in matching_materials]
    valid_placeholders = ",".join("?" for _ in valid_material_ids)
    db.execute(
        f"""
        UPDATE job_materials
        SET is_invoiced = 1,
            invoice_number = ?
        WHERE job_id = ?
            AND id IN ({valid_placeholders})
        """,
        [invoice_number, job_id, *valid_material_ids],
    )
    db.commit()
    flash(f"{len(valid_material_ids)} materiaalregels gekoppeld aan factuur {invoice_number}.", "success")
    return redirect(url_for("job_detail", job_id=job_id))


@app.post("/jobs/<int:job_id>/materials/<int:material_id>/update")
def update_job_material(job_id: int, material_id: int) -> Response:
    get_job_or_404(job_id)
    db = get_db()
    material = db.execute(
        "SELECT * FROM job_materials WHERE id = ? AND job_id = ?",
        (material_id, job_id),
    ).fetchone()
    if material is None:
        abort(404)

    product = get_product_or_404(material["product_id"])
    restore_material_allocations(
        db,
        material_id,
        material["product_id"],
        float(material["quantity"]),
        float(material["unit_cost_snapshot"]),
    )
    refresh_product_stock_state(db, material["product_id"])
    product = get_product_or_404(material["product_id"])
    quantity_raw = request.form.get("quantity", "").strip()
    meter_start_raw = request.form.get("meter_start", "").strip()
    meter_end_raw = request.form.get("meter_end", "").strip()
    meter_start = None
    meter_end = None

    if meter_start_raw or meter_end_raw:
        if not meter_start_raw or not meter_end_raw:
            flash("Vul beide kabelnummers in of laat ze leeg en geef handmatig het aantal meters op.", "error")
            return redirect(url_for("job_detail", job_id=job_id))
        try:
            meter_start = float(meter_start_raw.replace(",", "."))
            meter_end = float(meter_end_raw.replace(",", "."))
        except ValueError:
            flash("Kabelnummers moeten geldige getallen zijn.", "error")
            return redirect(url_for("job_detail", job_id=job_id))
        quantity = meter_end - meter_start
    else:
        try:
            quantity = parse_decimal(quantity_raw, "quantity used")
        except ValueError:
            flash("Het gebruikte aantal moet een geldig getal zijn.", "error")
            return redirect(url_for("job_detail", job_id=job_id))

    if quantity <= 0:
        if meter_start is not None and meter_end is not None:
            flash("Het eindnummer van de kabel moet groter zijn dan het startnummer.", "error")
        else:
            flash("Het gebruikte aantal moet groter zijn dan nul.", "error")
        return redirect(url_for("job_detail", job_id=job_id))

    quantity_delta = quantity - material["quantity"]
    if quantity_delta > product["stock_quantity"]:
        flash(
            f"Onvoldoende voorraad voor {product['description']}. Beschikbaar: {format_quantity(product['stock_quantity'])} {product['unit']}.",
            "error",
        )
        return redirect(url_for("job_detail", job_id=job_id))

    if request.form.get("clear_invoice") == "1":
        is_invoiced = 0
        invoice_number = ""
    else:
        is_invoiced = 1 if request.form.get("is_invoiced") == "on" else 0
        invoice_number = request.form.get("invoice_number", "").strip() if is_invoiced else ""
        if is_invoiced and not invoice_number:
            flash("Vul een factuurnummer in.", "error")
            return redirect(url_for("job_detail", job_id=job_id))

    try:
        allocations, weighted_unit_cost = consume_stock_fifo(db, material["product_id"], quantity)
    except ValueError:
        flash(
            f"Onvoldoende voorraad voor {product['description']}. Beschikbaar: {format_quantity(product['stock_quantity'])} {product['unit']}.",
            "error",
        )
        return redirect(url_for("job_detail", job_id=job_id))

    db.execute(
        """
        UPDATE job_materials
        SET quantity = ?, unit_cost_snapshot = ?, meter_start_snapshot = ?, meter_end_snapshot = ?,
            is_invoiced = ?, invoice_number = ?
        WHERE id = ? AND job_id = ?
        """,
        (
            quantity,
            weighted_unit_cost,
            meter_start,
            meter_end,
            is_invoiced,
            invoice_number,
            material_id,
            job_id,
        ),
    )
    for batch_id, allocation_quantity, allocation_unit_cost in allocations:
        db.execute(
            """
            INSERT INTO job_material_batch_allocations (
                job_material_id,
                batch_id,
                quantity,
                unit_cost_snapshot
            )
            VALUES (?, ?, ?, ?)
            """,
            (material_id, batch_id, allocation_quantity, allocation_unit_cost),
        )
    refresh_product_stock_state(db, material["product_id"])
    db.commit()
    flash("Materiaalregel bijgewerkt.", "success")
    return redirect(url_for("job_detail", job_id=job_id))


@app.post("/jobs/<int:job_id>/materials/<int:material_id>/delete")
def delete_job_material(job_id: int, material_id: int) -> Response:
    get_job_or_404(job_id)
    db = get_db()
    material = db.execute(
        "SELECT * FROM job_materials WHERE id = ? AND job_id = ?",
        (material_id, job_id),
    ).fetchone()
    if material is None:
        abort(404)
    restore_material_allocations(
        db,
        material_id,
        material["product_id"],
        float(material["quantity"]),
        float(material["unit_cost_snapshot"]),
    )
    refresh_product_stock_state(db, material["product_id"])
    db.execute("DELETE FROM job_materials WHERE id = ? AND job_id = ?", (material_id, job_id))
    db.commit()
    flash("Materiaal verwijderd en voorraad hersteld.", "success")
    return redirect(url_for("job_detail", job_id=job_id))


@app.get("/jobs/<int:job_id>/export.txt")
def export_job_text(job_id: int) -> Response:
    get_job_or_404(job_id)
    materials = get_db().execute(
        """
        SELECT *
        FROM job_materials
        WHERE job_id = ?
        ORDER BY id ASC
        """,
        (job_id,),
    ).fetchall()

    lines = [
        "Description - Price per piece - amount - total price",
    ]
    total = 0.0
    for material in materials:
        sale_price = calculate_sale_price(
            material["unit_cost_snapshot"], material["profit_margin_snapshot"]
        )
        line_total = round(material["quantity"] * sale_price, 2)
        total += line_total
        lines.append(
            f"{material['description_snapshot']} - "
            f"EUR {sale_price:.2f} - "
            f"{format_quantity(material['quantity'])} {material['unit_snapshot']} - "
            f"EUR {line_total:.2f}"
        )
    lines.extend(["", f"Total - EUR {total:.2f}"])
    return Response("\n".join(lines), mimetype="text/plain; charset=utf-8")


init_db()


if __name__ == "__main__":
    app.run(
        host=os.environ.get("FLASK_HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", "5000")),
        debug=os.environ.get("FLASK_DEBUG", "0") == "1",
    )
